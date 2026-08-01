#!/usr/bin/env python3
"""
108 Stitches — keeper board generator
-------------------------------------
Rosters come from ESPN (live source of truth for who's on each team now).
The Pre-Draft Worksheet is the overlay: keeper prices, price history, and how
each drafted player was acquired. Every current player is tagged:

    Kept    - drafted and on that team's keeper list (carries price history)
    Auction - drafted at auction, not kept (fresh price, no history)
    Traded  - acquired from another team via trade (inherits their sheet price)
    FA      - added from free agency mid-season (no auction value)

    python3 build_board.py

If ESPN can't be reached, it falls back to the draft-day rosters from the sheet.
"""
import unicodedata, re, json, sys
from pathlib import Path
from openpyxl import load_workbook

# ---- CONFIG -----------------------------------------------------------------
HERE      = Path(__file__).parent
WORKSHEET = HERE / "2026_Pre-Draft_Worksheet.xlsx"
TEMPLATE  = HERE / "template.html"
OUTPUT    = HERE / "108-stitches-keeper-board.html"
LEAGUE, SEASON, BUYIN = "108 Stitches", 2026, 290

USE_ESPN       = True
ESPN_LEAGUE_ID = 35759
ESPN_SEASON    = 2026
ESPN_S2        = None            # public league -> None
ESPN_SWID      = None            # public league -> None
ESPN_JSON_FILE = None            # optional: path to saved league JSON (offline/testing)
ESPN_URL_OVERRIDE = "https://espn-relay.baseball-gm.workers.dev/"   # your Cloudflare relay

# --- Live Auction data layer ---
PROJECTION_SEASON = 2026   # projections to show in the auction; flip to 2027 before the March 2027 auction
AUCTION_DIAG      = False   # raw player-sample dump (Phase 1 verify) — done
AUCTION_BAKE      = True    # build auction-players.json + a small on-page verification summary
KEEPER_DIAG       = True    # Phase 2: identify the live-ESPN field that marks a 2027 keeper
PAGES_URL      = "https://kokanjohn.github.io/108stitches/"   # used to reuse the last live snapshot if ESPN is down
OWNER_ALIAS    = {}              # {"ESPN Name": "Sheet Owner Name"} if a person's name differs

NAME_FIX = {"Jak Caglianone":"Jac Caglianone", "Sam Basallo":"Samuel Basallo",
            "Agustin Ramirez":"Agustín Ramírez", "Augustin Ramirez":"Agustín Ramírez"}
DRAFT_LABEL = {"Waller":"Justin Waller","T. Carter":"Travis Carter","Linthicum":"Brandon Linthicum",
 "Schottmiller":"Matt Schottmiller","Amos":"Jonathan Amos","Bonham":"Andrew Bonham",
 "Walker":"Eric Walker","Eanes":"Casey Eanes","Martinez":"Sam Martinez","Hines":"Matt Hines",
 "Moore":"John Moore","Brady":"Doug Brady","J. Carter":"Jeff Carter","Mitchell":"Chris Mitchell",
 "Cooper":"Scott Cooper","Stone":"Jonathan Stone"}
KEEPER_TABS = set(DRAFT_LABEL) | {"Travis","Brandon","Jeff"}  # tab names differ from Draft labels
KEEPER_TABS = {"Waller","Travis","Brandon","Schottmiller","Amos","Bonham","Walker","Brady",
 "Hines","Moore","Martinez","Jeff","Mitchell","Cooper","Stone","Eanes"}
# -----------------------------------------------------------------------------

def norm(s): return ("" if s is None else str(s)).replace("\xa0"," ").strip()
def keyof(s):
    s = unicodedata.normalize("NFKD", norm(s)).encode("ascii","ignore").decode().lower()
    return re.sub(r"[^a-z0-9]", "", s)
def toint(v):
    try: return int(v) if v not in (None,"") else None
    except (TypeError, ValueError): return None

def parse_worksheet(wb):
    """Return draft-day records (kept + auction), keyed index, and owner->team."""
    HIST, META, OWNER_TEAM = {}, {}, {}
    for i, r in enumerate(wb["Keeper Prices"].iter_rows(values_only=True)):
        if i == 0 or r is None: continue
        player = norm(r[2]) if len(r) > 2 else ""
        if not player: continue
        k = keyof(player)
        HIST[k] = {y: toint(r[idx]) for y, idx in zip((2022,2023,2024,2025), (5,6,7,8))}
        META[k] = {"name": NAME_FIX.get(player, player), "mlb": norm(r[3]) or "FA", "pos": norm(r[4])}
        owner, team = norm(r[1]), norm(r[0])
        if owner and owner not in OWNER_TEAM: OWNER_TEAM[owner] = team

    KEPT, KEEP_PRICE = set(), {}
    for tab in KEEPER_TABS:
        for r in list(wb[tab].iter_rows(values_only=True))[3:]:
            if r is None: continue
            name = norm(r[1]) if len(r) > 1 else ""
            if name and name.lower() != "name":
                KEPT.add(keyof(name)); KEEP_PRICE[keyof(name)] = toint(r[2]) if len(r) > 2 else None

    rows = list(wb["Draft"].iter_rows(values_only=True))
    cols = {c: DRAFT_LABEL[norm(rows[2][c])] for c in range(1, len(rows[2]), 2)
            if norm(rows[2][c]) in DRAFT_LABEL}
    records, index = [], {}
    for r in rows[6:32]:
        if r is None: continue
        slot = norm(r[0])
        for c, owner in cols.items():
            if c >= len(r): continue
            name = norm(r[c]); price = toint(r[c+1]) if c+1 < len(r) else None
            if not name or name.upper() == "FILL IN" or name.startswith("#"): continue
            k = keyof(name); kept = k in KEPT; m = META.get(k)
            disp = m["name"] if m else NAME_FIX.get(name, name)
            cur = price if price is not None else (KEEP_PRICE.get(k) if kept else None)
            hist = HIST.get(k, {}) if kept else {}
            rec = {"team": OWNER_TEAM.get(owner, owner), "owner": owner, "player": disp,
                   "mlb": (m["mlb"] if m else "FA"), "pos": slot, "kept": kept,
                   "tag": "kept" if kept else "auction",
                   "p": {**{y: (hist.get(y) if kept else None) for y in (2022,2023,2024,2025)}, 2026: cur}}
            records.append(rec)
            index[k] = rec
    return records, index, OWNER_TEAM

def classify(acq, sp, owner):
    a = acq or ""
    if a == "TRADE": return "traded"
    if a in ("ADD", "WAIVER", "FREEAGENT", "ADD_WAIVER"): return "fa"
    if a in ("DRAFT", "DRAFTED"):
        if sp and keyof(sp["owner"]) == keyof(owner): return "kept" if sp["kept"] else "auction"
        if sp: return "traded"          # drafted flag but sheet shows another owner -> treated as trade
        return "auction"                # drafted but not on the sheet
    if sp: return "kept" if sp["kept"] else "auction"
    return "fa"

def roster_rank(e, idx):
    """ESPN roster display order: active batters C->UT, bench/IL batters,
       active pitchers P/SP/RP, bench/IL pitchers."""
    BAT = {0:0, 1:1, 2:2, 3:3, 4:4, 6:5, 7:6, 5:7, 8:7, 9:7, 10:7, 11:8, 12:8}
    PIT = {13:0, 14:1, 15:2}
    s = e.get("slot_id")
    if s in (13, 14, 15):
        return (2, PIT.get(s, 9), idx)
    if s in (16, 17):
        sub = 0 if s == 16 else 1          # bench before IL
        return ((3 if e.get("is_pitcher") else 1), sub, idx)
    return (0, BAT.get(s, 9), idx)          # active batter

def build_from_espn(rosters, index, OWNER_TEAM):
    from espn_live import key as ekey, fuzzy
    keys = list(index.keys())
    # order every roster spot the way ESPN displays it, within each team
    ordered = sorted(enumerate(rosters),
                     key=lambda t: ((t[1]["team"] or t[1]["owner"] or ""), roster_rank(t[1], t[0])))
    records, teams, matched = [], {}, 0
    for _, e in ordered:
        owner, name, acq = e["owner"], e["player"], e["acq"]
        k = ekey(name); sp = index.get(k)
        if sp is None:
            nk = fuzzy(k, keys)
            if nk: sp = index.get(nk)
        if sp: matched += 1
        tag = classify(acq, sp, owner)
        team = e["team"] or OWNER_TEAM.get(owner, owner)     # ESPN's current team name wins
        s = e.get("slot_id"); is_pit = e.get("is_pitcher")
        grp = "pit" if (s in (13,14,15) or (s in (16,17) and is_pit)) else "bat"
        p = {**{y: (sp["p"][y] if sp else None) for y in (2022,2023,2024,2025)},
             2026: (sp["p"][2026] if sp else None)}
        rec = {"team": team, "owner": owner,
               "player": (sp["player"] if sp else name),
               "mlb": (sp["mlb"] if sp else (e["mlb"] or "FA")),
               "pos": (e["pos"] or (sp["pos"] if sp else "")),  # ESPN's current roster slot
               "elig": e.get("elig", ""), "grp": grp,
               "kept": tag == "kept", "tag": tag, "p": p}
        records.append(rec)
        t = teams.setdefault(team, {"team": team, "owner": owner, "count": 0, "kept": 0, "total2026": 0})
        t["count"] += 1; t["kept"] += 1 if tag == "kept" else 0
        if p[2026]: t["total2026"] += p[2026]
    return records, teams, matched

def team_meta(records):
    teams = {}
    for r in records:
        t = teams.setdefault(r["team"], {"team": r["team"], "owner": r["owner"], "count": 0, "kept": 0, "total2026": 0})
        t["count"] += 1; t["kept"] += 1 if r["tag"] == "kept" else 0
        if r["p"][2026]: t["total2026"] += r["p"][2026]
    return teams

def compute_standings(league):
    """Top-3 teams per H2H category + total 'Moves' for the $1/move pool.
    Stat IDs verified against this league's actual standings."""
    import re as _re
    CAT_STATID = {"R":20, "HR":5, "RBI":21, "SB":23, "AVG":2, "OPS":18,
                  "K":48, "QS":63, "BAA":38, "ERA":47, "WHIP":41, "SVHD":83}
    CAT_ORDER = ["R","HR","RBI","SB","AVG","OPS","K","QS","BAA","ERA","WHIP","SVHD"]
    REV = {"BAA","ERA","WHIP"}
    def tname(t):
        n = (t.get("name") or " ".join(x for x in (t.get("location"), t.get("nickname")) if x)
             or f"Team {t.get('id')}")
        return _re.sub(r"\s+", " ", n).strip()
    teams = league.get("teams") or []
    tds = []
    for t in teams:
        tc = t.get("transactionCounter") or {}
        tds.append({"team": tname(t), "vals": t.get("valuesByStat") or {},
                    "acq": tc.get("acquisitions") or 0})
    total_moves = sum(td["acq"] for td in tds)
    slice_ = total_moves / len(CAT_ORDER) if CAT_ORDER else 0
    cats = []
    earn, leads = {}, {}
    for label in CAT_ORDER:
        sid = CAT_STATID[label]; rev = label in REV
        rows = []
        for td in tds:
            v = td["vals"].get(str(sid))
            if v is None: v = td["vals"].get(sid)
            if v is not None:
                try: rows.append({"team": td["team"], "value": float(v)})
                except (TypeError, ValueError): pass
        rows.sort(key=lambda r: r["value"], reverse=not rev)
        cats.append({"label": label, "statId": sid, "reverse": rev, "top3": rows[:3]})
        if rows:
            best = rows[0]["value"]
            winners = [r["team"] for r in rows if r["value"] == best]   # ties split the slice
            share = slice_ / len(winners)
            for w in winners:
                earn[w] = earn.get(w, 0) + share
                leads[w] = leads.get(w, 0) + 1
    payout = sorted([{"team": t, "earned": round(earn[t]), "cats": leads[t]} for t in earn],
                    key=lambda x: (-x["earned"], x["team"]))
    moves = sorted([{"team": td["team"], "moves": td["acq"]} for td in tds],
                   key=lambda x: (-x["moves"], x["team"]))
    return {"pool": total_moves, "totalMoves": total_moves, "slice": round(slice_, 2),
            "moves": moves, "categories": cats, "payout": payout}

def auction_diagnostic():
    """Phase 1: pull a few players from the ESPN pool via the relay and return a
    trimmed sample of their projection/history structure, to design the bake against."""
    if not (USE_ESPN and ESPN_URL_OVERRIDE):
        return {"ok": False, "note": "relay not configured"}
    base = ESPN_URL_OVERRIDE.rstrip("/")
    sep = "&" if "?" in base else "?"
    url = f"{base}{sep}mode=players&season={PROJECTION_SEASON}&limit=8"
    try:
        from espn_live import fetch_league
        js = fetch_league(ESPN_LEAGUE_ID, PROJECTION_SEASON, url=url)
    except Exception as e:
        return {"ok": False, "note": f"{type(e).__name__}: {e}", "url": url}
    players = js.get("players") if isinstance(js, dict) else None
    if not players:
        keys = list(js.keys())[:20] if isinstance(js, dict) else str(type(js))
        return {"ok": False, "note": "no 'players' array in response", "topKeys": keys, "url": url}
    sample = []
    for item in players[:5]:
        p = item.get("player") or item
        stats = []
        for s in (p.get("stats") or []):
            entry = {k: s.get(k) for k in ("seasonId", "statSourceId", "statSplitTypeId", "appliedTotal")}
            if s.get("statSplitTypeId") == 0:          # season totals only (skip weekly splits)
                entry["stats"] = s.get("stats")
            stats.append(entry)
        sample.append({
            "id": p.get("id"), "fullName": p.get("fullName"),
            "proTeamId": p.get("proTeamId"), "defaultPositionId": p.get("defaultPositionId"),
            "eligibleSlots": p.get("eligibleSlots"), "stats": stats,
        })
    return {"ok": True, "season": PROJECTION_SEASON, "poolCount": len(players), "sample": sample, "url": url}

# stat-id maps for the auction display (verified against the ESPN payload)
HIT_CATS = {"R": "20", "HR": "5", "RBI": "21", "SB": "23", "AVG": "2", "OPS": "18"}
PIT_CATS = {"K": "48", "QS": "63", "BAA": "38", "ERA": "47", "WHIP": "41", "SVHD": "83"}
SV_HLD_CANDIDATES = ["57", "58", "59", "60", "83"]   # to pin down SVHD from a real reliever

def _cat_line(stats_obj, catmap, want_hit, want_pit):
    d = {}
    if not stats_obj:
        return d
    if want_hit:
        for name, sid in HIT_CATS.items():
            d[name] = stats_obj.get(sid)
    if want_pit:
        for name, sid in PIT_CATS.items():
            d[name] = stats_obj.get(sid)
        if d.get("SVHD") is None:                     # fallback if id 83 is absent
            sv, hl = stats_obj.get("57"), stats_obj.get("60")
            if sv is not None or hl is not None:
                d["SVHD"] = (sv or 0) + (hl or 0)
    return d

def build_auction_players(limit=1500):
    """Phase 1 bake: pull the ESPN player pool and produce compact records for the
    auction display — id, name, MLB team, positions, PROJECTION_SEASON projections,
    and the previous season's actuals. Returns (full_dict, check_dict)."""
    from espn_live import fetch_league, eligibility, PRO_TEAM
    if not (USE_ESPN and ESPN_URL_OVERRIDE):
        return None, {"ok": False, "note": "relay not configured"}
    base = ESPN_URL_OVERRIDE.rstrip("/"); sep = "&" if "?" in base else "?"
    url = f"{base}{sep}mode=players&season={PROJECTION_SEASON}&limit={limit}"
    try:
        js = fetch_league(ESPN_LEAGUE_ID, PROJECTION_SEASON, url=url, timeout=60)
    except Exception as e:
        return None, {"ok": False, "note": f"{type(e).__name__}: {e}", "url": url}
    raw = js.get("players") if isinstance(js, dict) else None
    if not raw:
        return None, {"ok": False, "note": "no players array", "url": url}

    proj_season, prev_season = PROJECTION_SEASON, PROJECTION_SEASON - 1
    players = []
    probe = {k: {"n": 0, "max": 0, "top": []} for k in SV_HLD_CANDIDATES}
    for item in raw:
        p = item.get("player") or item
        pid, name = p.get("id"), p.get("fullName")
        if not pid or not name:
            continue
        elig = p.get("eligibleSlots") or []
        pos, is_pitcher = eligibility(elig)
        is_two_way = (not is_pitcher) and any(i in (13, 14, 15) for i in elig)
        proj = prev = None
        for s in (p.get("stats") or []):
            if s.get("statSplitTypeId") != 0:
                continue
            ssn, src = s.get("seasonId"), s.get("statSourceId")
            if ssn == proj_season and src == 1:
                proj = s.get("stats") or {}
            elif ssn == prev_season and src == 0:
                prev = s.get("stats") or {}
        want_hit = (not is_pitcher) or is_two_way
        want_pit = is_pitcher or is_two_way
        players.append({
            "id": pid, "name": name, "mlb": PRO_TEAM.get(p.get("proTeamId"), ""),
            "pos": pos, "pitcher": is_pitcher, "twoWay": is_two_way, "rookie": prev is None,
            "proj": _cat_line(proj, None, want_hit, want_pit),
            "prev": _cat_line(prev, None, want_hit, want_pit),
            "prevSeason": prev_season,
        })
        if is_pitcher and proj:
            for k in SV_HLD_CANDIDATES:
                v = proj.get(k)
                if isinstance(v, (int, float)) and v:
                    probe[k]["n"] += 1
                    probe[k]["max"] = max(probe[k]["max"], v)
                    probe[k]["top"].append([name, v])
    for k in probe:
        probe[k]["top"] = sorted(probe[k]["top"], key=lambda x: -x[1])[:4]

    full = {"ok": True, "season": proj_season, "prevSeason": prev_season,
            "count": len(players), "players": players}
    # small on-page verification summary (a few representative records + the SVHD probe)
    def pick(f):
        return next((r for r in players if f(r)), None)
    samples = [r for r in [pick(lambda r: not r["pitcher"] and not r["twoWay"]),
                           pick(lambda r: r["pitcher"]),
                           pick(lambda r: r["twoWay"]),
                           pick(lambda r: r["rookie"])] if r]
    check = {"ok": True, "season": proj_season, "prevSeason": prev_season,
             "count": len(players),
             "hitters": sum(1 for r in players if not r["pitcher"]),
             "pitchers": sum(1 for r in players if r["pitcher"]),
             "twoWay": sum(1 for r in players if r["twoWay"]),
             "rookies": sum(1 for r in players if r["rookie"]),
             "svhdProbe": probe, "samples": samples}
    return full, check

def keeper_diagnostic(league):
    """Find the live-ESPN field that marks a 2027 keeper. Dumps candidate fields per team
    plus full detail for the two teams named in testing (Iguchigang, NATy Lights)."""
    members = {}
    for m in league.get("members", []):
        full = f"{(m.get('firstName') or '').strip()} {(m.get('lastName') or '').strip()}".strip()
        members[m.get("id")] = full or (m.get("displayName") or "")
    entry_keys, pool_keys = set(), set()
    summary, detail = [], {}
    want = ("iguchi", "naty")
    for t in league.get("teams", []):
        tname = (t.get("name") or " ".join(x for x in (t.get("location"), t.get("nickname")) if x) or "").strip()
        owners = t.get("owners") or []
        owner = members.get(owners[0], "") if owners else ""
        rows = (t.get("roster") or {}).get("entries", [])
        kv = kvf = keepflag = 0
        det = []
        for e in rows:
            entry_keys.update(e.keys())
            ppe = e.get("playerPoolEntry") or {}
            pool_keys.update(ppe.keys())
            pl = ppe.get("player") or {}
            v_kv, v_kvf = ppe.get("keeperValue"), ppe.get("keeperValueFuture")
            keepish = {k: v for k, v in list(e.items()) + list(ppe.items()) if "keep" in k.lower()}
            if v_kv not in (None, 0): kv += 1
            if v_kvf not in (None, 0): kvf += 1
            if any(v is True for v in keepish.values()): keepflag += 1
            if any(w in tname.lower() for w in want):
                det.append({"player": pl.get("fullName"), "acq": e.get("acquisitionType"),
                            "keeperValue": v_kv, "keeperValueFuture": v_kvf, "keepish": keepish})
        summary.append({"team": tname, "owner": owner, "entries": len(rows),
                        "keeperValueSet": kv, "keeperValueFutureSet": kvf, "keepFlagTrue": keepflag})
        if det:
            detail[tname] = det
    return {"entryKeys": sorted(entry_keys), "poolKeys": sorted(pool_keys),
            "summary": sorted(summary, key=lambda x: x["team"]), "detail": detail}

def build():
    if not WORKSHEET.exists(): sys.exit(f"Can't find the worksheet: {WORKSHEET}")
    if not TEMPLATE.exists():  sys.exit("Can't find template.html next to this script.")
    wb = load_workbook(WORKSHEET, read_only=True, data_only=True)
    draft_records, index, OWNER_TEAM = parse_worksheet(wb)

    from datetime import datetime, timezone
    try:
        from zoneinfo import ZoneInfo
        built_at = datetime.now(ZoneInfo("America/New_York")).strftime("%b %d, %Y · %I:%M %p %Z")
    except Exception:
        built_at = datetime.now(timezone.utc).strftime("%b %d, %Y · %H:%M UTC")
    records, teams, live = draft_records, team_meta(draft_records), False
    stale = False; snapshot_at = ""; standings = None
    live_error = live_hint = target = ""
    keeper_diag = None
    if USE_ESPN:
        target = ESPN_URL_OVERRIDE or (
            f"https://lm-api-reads.fantasy.espn.com/apis/v3/games/flb/seasons/{ESPN_SEASON}"
            f"/segments/0/leagues/{ESPN_LEAGUE_ID}?view=mTeam&view=mRoster")
        try:
            import urllib.error
            from espn_live import fetch_league, current_rosters
            raw = Path(ESPN_JSON_FILE).read_text(encoding="utf-8") if ESPN_JSON_FILE else None
            league = fetch_league(ESPN_LEAGUE_ID, ESPN_SEASON, ESPN_S2, ESPN_SWID,
                                  local_json=raw, url=(ESPN_URL_OVERRIDE or None))
            rosters = current_rosters(league, OWNER_ALIAS)
            records, teams, matched = build_from_espn(rosters, index, OWNER_TEAM)
            standings = compute_standings(league)
            live = True
            if KEEPER_DIAG:
                try: keeper_diag = keeper_diagnostic(league)
                except Exception as e: keeper_diag = {"note": f"{type(e).__name__}: {e}"}
            counts = {t: sum(1 for r in records if r["tag"] == t) for t in ("kept","auction","traded","fa")}
            print(f"  live: {len(records)} current roster spots — {counts['kept']} kept, "
                  f"{counts['auction']} auction, {counts['traded']} traded, {counts['fa']} free agents "
                  f"({matched} matched to sheet prices)")
            print(f"  standings: {len(standings['categories'])} categories, "
                  f"{standings['totalMoves']} total moves (${standings['pool']} pool)")
        except Exception as e:
            live_error = f"{type(e).__name__}: {e}"
            if isinstance(e, urllib.error.HTTPError):
                body = ""
                try: body = e.read().decode("utf-8", "replace")
                except Exception: pass
                msg = ""
                if body:
                    try: msg = (json.loads(body).get("messages") or [""])[0]
                    except Exception: msg = body[:160]
                live_error = f"HTTP {e.code} from ESPN" + (f' — "{msg}"' if msg else "")
                if e.code in (401, 403):
                    live_hint = ("ESPN refused the request. If the league is private, the relay's espn_s2 "
                                 "cookie is missing or expired — refresh it in the Cloudflare Worker and Deploy. "
                                 "If you switched the league to public, double-check that setting saved.")
            elif isinstance(e, urllib.error.URLError):
                live_hint = "Couldn't reach the URL below — check the relay address is right and deployed."
            # keep the last live data: reuse the previously published page instead of the sheet
            try:
                import urllib.request
                with urllib.request.urlopen(PAGES_URL, timeout=20) as resp:
                    prev_html = resp.read().decode("utf-8", "replace")
                a = prev_html.index("const DATA = ") + len("const DATA = ")
                b = prev_html.index(";\nconst $ = s", a)
                prev = json.loads(prev_html[a:b])
                if prev.get("records"):
                    records, teams = prev["records"], prev["teams"]
                    standings = prev.get("standings")
                    stale = True
                    snapshot_at = prev.get("snapshot_at") or prev.get("built_at", "")
                    print(f"  live: ESPN unavailable ({live_error}) — reused last live snapshot from {snapshot_at}")
            except Exception as e2:
                print(f"  live: ESPN unavailable ({live_error}); no snapshot ({type(e2).__name__}) — using the sheet")

    # 2027 keeper cost (skip for a reused snapshot — those records already carry it).
    # No auction price on the sheet -> flat $5 to keep next year. A genuine escalation only
    # runs when 2026 is actually ABOVE 2025 (prices never fall when kept), giving
    # 2*2026 - 2025 + 1. If 2026 is not above 2025, the player was re-auctioned/re-priced and
    # the old history is stale: clear it and treat 2026 as a fresh base (+2). This also covers
    # a freshly-auctioned player who simply has no 2025 on record.
    if not stale:
        for r in records:
            p26, p25 = r["p"].get(2026), r["p"].get(2025)
            if p26 is None:
                r["p"][2027] = 5
            elif p25 is not None and p26 > p25:
                r["p"][2027] = 2*p26 - p25 + 1
            else:
                if p25 is not None:                       # stale history above a lower price -> reset
                    for y in (2022, 2023, 2024, 2025):
                        r["p"][y] = None
                r["p"][2027] = p26 + 2

    teams_out = teams if isinstance(teams, list) else sorted(teams.values(), key=lambda x: -x["total2026"])
    data = {"league": LEAGUE, "season": SEASON, "buyin": BUYIN, "live": live,
            "stale": stale, "snapshot_at": snapshot_at, "standings": standings,
            "built_at": built_at, "live_error": live_error, "live_hint": live_hint,
            "live_target": target, "via_relay": bool(ESPN_URL_OVERRIDE),
            "records": records, "teams": teams_out}
    if keeper_diag is not None:
        data["keeperDiag"] = keeper_diag
    if AUCTION_DIAG:
        try: data["auctionDiag"] = auction_diagnostic()
        except Exception as e: data["auctionDiag"] = {"ok": False, "note": f"{type(e).__name__}: {e}"}
    if AUCTION_BAKE:
        try:
            full, check = build_auction_players()
            # per-team budget inputs: keeper count + committed 2027 salary (kept players)
            tmap = {t["team"]: {"team": t["team"], "owner": t.get("owner", ""),
                                "keeperCount": 0, "keeperSalary2027": 0} for t in teams_out}
            for r in records:
                if r.get("kept") and r["team"] in tmap:
                    tmap[r["team"]]["keeperCount"] += 1
                    s = r["p"].get(2027)
                    if s: tmap[r["team"]]["keeperSalary2027"] += s
            teams_budget = sorted(tmap.values(), key=lambda x: x["team"])
            check["teams"] = teams_budget
            data["auctionBakeCheck"] = check
            if full is not None:
                full["teams"] = teams_budget
                with open("auction-players.json", "w", encoding="utf-8") as fh:
                    json.dump(full, fh, ensure_ascii=False)
                print(f"✓ Wrote auction-players.json — {full['count']} players, {len(teams_budget)} teams")
        except Exception as e:
            data["auctionBakeCheck"] = {"ok": False, "note": f"{type(e).__name__}: {e}"}
    html = TEMPLATE.read_text(encoding="utf-8").replace("/*__DATA__*/", json.dumps(data, ensure_ascii=False))
    OUTPUT.write_text(html, encoding="utf-8")
    src = ("ESPN live rosters" if live else
           (f"last live snapshot ({snapshot_at})" if stale else "draft-day rosters (ESPN off)"))
    print(f"✓ Built {OUTPUT.name} — {len(records)} players across {len(teams_out)} teams · {src}")

if __name__ == "__main__":
    build()
